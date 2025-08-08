import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def format():
    # Define your date columns
    date_columns = [
        "Travel Start Date",
        "Travel End Date",
        "First Submitted Date",
        "Last Submitted Date",
        "Reports to Approval 2",
        "Budget Approval",
        "Approved Date / Sent for Payment Date",
        "Transaction Date",
        "Processor Approval Date",
        "Sent for Payment Date",
        "Paid Date"
    ]

    # STEP 1: Load the Excel file
    input_file = 'data/CombineAndFormatData/master_expense_report.xlsx'
    df = pd.read_excel(input_file)

    # STEP 2: Remove time from specified date columns (if present)
    for col in date_columns:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.date

    # STEP 3: Save to Excel with formatting
    output_file = 'data/CombineAndFormatData/master_expense_report_formatted.xlsx'
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')

        # Access workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Apply date format to appropriate columns
        for idx, col in enumerate(df.columns, start=1):  # start=1 because Excel columns are 1-based
            if col in date_columns:
                col_letter = get_column_letter(idx)
                for cell in worksheet[col_letter][1:]:  # skip header row
                    cell.number_format = 'mm/dd/yy'

    print("Done Formatting Dates")