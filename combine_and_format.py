import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def load_excel_file_task(file_info):
    """Load a single Excel file - used for parallel processing"""
    path, sheet_name, header, columns, name = file_info
    df = pd.read_excel(path, sheet_name=sheet_name, header=header)
    if columns:
        df = df[columns]
    return name, df

def combine_and_format(expense_etd_path=None, ee_active_path=None, expense_cf_path=None, expense_ppsa_path=None, request_rit_path=None):
    from concurrent.futures import ProcessPoolExecutor
    
    # Define file loading tasks
    file_tasks = [
        (expense_etd_path or "./data/CombineAndFormatData/Expense - Expense Type Detail (SLO) 2024.xlsx", "Details_1", 8, None, "expense_etd"),
        (ee_active_path or "./data/CombineAndFormatData/EE Active.xlsx", "EE-Active", 0, ["Emplid", "Empl Status Pay Ldescr", "Division", "Deptid Ldescr", "Position Ldescr"], "employee_active"),
        (expense_cf_path or "./data/CombineAndFormatData/Expense - Expense Reports with CF Information and Comments (2).xlsx", "Summary_1", 6, ["Report Key", "Request ID and Destination", "Total Approved Amount (rpt)", "Processor Approval Date"], "expense_cf"),
        (expense_ppsa_path or "./data/CombineAndFormatData/Expense - Processor Paid Summary Account.xlsx", "Page1_1", 4, ["Sent for Payment Date", "Paid Date", "Report Key", "Transaction Date", "Expense Type", "Approved Amount"], "expense_ppsa"),
        (request_rit_path or "./data/CombineAndFormatData/Request - Risk International Travel with Header Comments (1).xlsx", "Request - Risk International_1", 7, ["Request ID", "Authorized Date", "Destination City/Location", "Destination Country"], "request_rit")
    ]
    
    # Load all files in parallel using processes
    dataframes = {}
    with ProcessPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(load_excel_file_task, task) for task in file_tasks]
        for future in futures:
            name, df = future.result()
            dataframes[name] = df
    
    # Rename column for employee_active
    dataframes["employee_active"].rename(columns={"Empl Status Pay Ldescr": "Active/Term Date"}, inplace=True)
    # Perform merges using loaded dataframes
    merged_df = pd.merge(
        dataframes["expense_etd"],
        dataframes["employee_active"],
        left_on="Employee ID",
        right_on="Emplid",
        how="left"
    )
    merged_df.drop(columns=["Emplid"], inplace=True)

    merged_df = pd.merge(
        merged_df,
        dataframes["expense_cf"],
        on="Report Key",
        how="left"
    )

    merged_df = pd.merge(
        merged_df,
        dataframes["expense_ppsa"],
        left_on=["Report Key", "Transaction Date", "Expense Type", "Approved Amount (rpt)"],
        right_on=["Report Key", "Transaction Date", "Expense Type", "Approved Amount"],
        how="left"
    )
    merged_df.drop(columns=["Approved Amount"], inplace=True)

    merged_df = pd.merge(
        merged_df,
        dataframes["request_rit"],
        left_on=["Request ID(s)"],
        right_on=["Request ID"],
        how="left"
    )
    merged_df.drop(columns=["Request ID"], inplace=True)

    # Create audit_reports directory if it doesn't exist
    os.makedirs("audit_reports", exist_ok=True)
    
    # Create formatted Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Expenses Report"
    
    # Add data to worksheet
    for r in dataframe_to_rows(merged_df, index=False, header=True):
        ws.append(r)
    
    # Format header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Return the merged DataFrame instead of saving
    return merged_df