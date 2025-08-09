# services/report_writer.py
from pathlib import Path
from datetime import datetime
from typing import Optional, Union, Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from config.settings import REPORTS_DIR


def flag_audit_rows(
    df_original: pd.DataFrame,
    df_clean: pd.DataFrame,               # kept for signature parity (not used here)
    violation_rows: list,
    exception_rows: list
) -> pd.DataFrame:
    """
    Flags rows in df_original based on violation/exception row numbers.
    Adds 'Audit Flag' with 'Violation'/'Exception'/''.
    IMPORTANT: We flag against df_original['Original Row'] (not df_clean).
    """
    def get_flag(row_number: int) -> str:
        if row_number in violation_rows:
            return "Violation"
        if row_number in exception_rows:
            return "Exception"
        return ""

    df_out = df_original.copy()
    if "Original Row" not in df_out.columns:
        raise KeyError("Missing required column 'Original Row' in df_original")

    df_out["Audit Flag"] = df_out["Original Row"].apply(get_flag)
    return df_out


def save_to_excel_with_formatting(
    df_flagged: pd.DataFrame,
    output_path: Optional[Union[str, Path]] = None
) -> str:
    """
    Saves df_flagged to Excel with row color fills based on 'Audit Flag'.
    Returns the output path as string.
    """
    # Columns to render as dates
    date_columns = [
        "Travel Start Date", "Travel End Date", "First Submitted Date", "Last Submitted Date",
        "Reports to Approval 2", "Budget Approval", "Approved Date / Sent for Payment Date",
        "Transaction Date", "Processor Approval Date", "Sent for Payment Date", "Paid Date"
    ]

    # Default output path (timestamped) under project_root/audit_reports
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = REPORTS_DIR / f"Audited_Expenses_{timestamp}.xlsx"
    else:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

    # Coerce dates for display
    df_x = df_flagged.copy()
    for col in date_columns:
        if col in df_x.columns:
            df_x[col] = pd.to_datetime(df_x[col], errors='coerce').dt.date

    wb = Workbook()
    ws = wb.active
    ws.title = "Audited Data"

    for r in dataframe_to_rows(df_x, index=False, header=True):
        ws.append(r)

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

    if "Audit Flag" not in df_x.columns:
        raise KeyError("Missing required column 'Audit Flag' in df_flagged")
    audit_flag_col = list(df_x.columns).index("Audit Flag") + 1  # 1-based

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # skip header
        cell = row[audit_flag_col - 1]
        if cell.value == "Violation":
            for c in row:
                c.fill = red_fill
        elif cell.value == "Exception":
            for c in row:
                c.fill = yellow_fill

    # Excel display format for dates
    for idx, col in enumerate(df_x.columns, start=1):
        if col in date_columns:
            col_letter = get_column_letter(idx)
            for cell in ws[col_letter][1:]:
                cell.number_format = "mm/dd/yy"

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"✅ Saved audited file to: {output_path}")
    return str(output_path)


def create_violations_exceptions_report(
    df_flagged: pd.DataFrame,
    audit_results: Optional[list] = None
) -> Dict[str, Optional[str]]:
    """
    Creates separate Violations and Exceptions workbooks.
    Returns dict of written paths: {"violations": str|None, "exceptions": str|None}
    """
    paths: Dict[str, Optional[str]] = {"violations": None, "exceptions": None}
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Safely access "Audit Flag"
    flag_col = "Audit Flag"
    if flag_col not in df_flagged.columns:
        raise KeyError("Missing required column 'Audit Flag' in df_flagged")

    violations_df = df_flagged[df_flagged[flag_col] == "Violation"].copy()
    exceptions_df = df_flagged[df_flagged[flag_col] == "Exception"].copy()
    print(f"Found {len(violations_df)} violations and {len(exceptions_df)} exceptions")

    def _write(df: pd.DataFrame, out_path: Path, fill: PatternFill, title: str) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = title
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.fill = fill
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        print(f"✅ Saved {title} report to: {out_path}")
        return str(out_path)

    if not violations_df.empty:
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        v_path = REPORTS_DIR / f"Violations_Report_{timestamp}.xlsx"
        paths["violations"] = _write(violations_df, v_path, red, "Violations")

    if not exceptions_df.empty:
        yellow = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
        e_path = REPORTS_DIR / f"Exceptions_Report_{timestamp}.xlsx"
        paths["exceptions"] = _write(exceptions_df, e_path, yellow, "Exceptions")

    return paths


def audit_and_flag(
    df_original: pd.DataFrame,
    df_clean: pd.DataFrame,
    bedrock_runtime
):
    """
    Runs the LLM audit for a sample of employee-report groups, flags df_original,
    saves the audited file and split reports, and returns (audited_subset, paths_dict).
    NOTE: This is long-running; normally you'd keep it in a controller, but provided
    here since you said you aren't using controllers right now.
    """
    # Import here to avoid circular imports
    from services.auditor import run_audit_for_multiple_employees

    # Run audit via Bedrock (sampled groups inside the function)
    violation_rows, exception_rows, audit_results = run_audit_for_multiple_employees(
        df_clean, bedrock_runtime
    )

    # Basic sanity checks
    df_o = df_original.copy()
    df_o.columns = df_o.columns.str.strip()
    for col in ("Employee ID", "Report Key", "Original Row"):
        if col not in df_o.columns:
            raise KeyError(
                f"❌ Required column '{col}' not found in df_original.\n"
                f"Available: {list(df_o.columns)}"
            )

    # Build set of all rows that belong to the audited groups (even if not flagged)
    audited_row_numbers = set(violation_rows + exception_rows)
    for r in audit_results:
        emp_id = str(r.get("employee_id"))
        report_key = str(r.get("report_key"))
        mask = (df_o["Employee ID"].astype(str) == emp_id) & \
               (df_o["Report Key"].astype(str) == report_key)
        grp = df_o[mask]
        if grp.empty:
            print(f"⚠️ No matching rows for Employee ID: {emp_id}, Report Key: {report_key}")
        audited_row_numbers.update(grp["Original Row"].tolist())

    # Keep only audited rows; then flag
    audited_subset = df_o[df_o["Original Row"].isin(audited_row_numbers)].copy()
    audited_subset = flag_audit_rows(audited_subset, df_clean, violation_rows, exception_rows)

    # Write files and return paths
    audited_path = save_to_excel_with_formatting(audited_subset)
    split_paths = create_violations_exceptions_report(audited_subset, audit_results)

    return audited_subset, {
        "audited_excel": audited_path,
        "violations": split_paths.get("violations"),
        "exceptions": split_paths.get("exceptions"),
    }
