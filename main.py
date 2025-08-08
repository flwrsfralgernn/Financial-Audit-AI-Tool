import boto3
import json
import random
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from combine_and_format import combine_and_format
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import pandas as pd
from boto3 import Session
from config import aws_access_key_id, aws_secret_access_key, aws_session_token


def load_excel_file(file_path):
    xls = pd.ExcelFile(file_path)
    df = pd.read_excel(xls, sheet_name=xls.sheet_names[0])
    return df

def audit_and_flag(df_original, df_clean, bedrock_runtime):
    # Run audit
    violation_rows, exception_rows, audit_results = run_audit_for_multiple_employees(df_clean, bedrock_runtime)

    # Ensure column names are stripped of spaces
    df_original.columns = df_original.columns.str.strip()
    audited_row_numbers = set(violation_rows + exception_rows)

    # Debug: print columns to check for header issues
    required_columns = ['Employee ID', 'Report Key', 'Original Row']
    for col in required_columns:
        if col not in df_original.columns:
            raise KeyError(f"❌ Required column '{col}' not found in df_original.\nAvailable columns: {df_original.columns.tolist()}")

    # Add remaining rows from the groups that were checked
    for r in audit_results:
        emp_id = r['employee_id']
        report_key = r['report_key']

        # Ensure matching types
        mask = (
            df_original['Employee ID'].astype(str) == str(emp_id)
        ) & (
            df_original['Report Key'].astype(str) == str(report_key)
        )

        group_rows = df_original[mask]
        if group_rows.empty:
            print(f" No matching rows found for Employee ID: {emp_id}, Report Key: {report_key}")

        audited_row_numbers.update(group_rows['Original Row'].tolist())

    # Filter and flag
    audited_subset = df_original[df_original["Original Row"].isin(audited_row_numbers)].copy()

    def get_flag(row_number):
        if row_number in violation_rows:
            return "Violation"
        elif row_number in exception_rows:
            return "Exception"
        return ""

    audited_subset["Audit Flag"] = audited_subset["Original Row"].apply(get_flag)

    save_to_excel_with_formatting(audited_subset)

    return audited_subset




def invoke_claude_model(prompt: str, bedrock_runtime) -> str:
    """
    Sends a prompt to Claude 3 Sonnet via Amazon Bedrock and returns the full streamed response text.
    """

    response = bedrock_runtime.invoke_model_with_response_stream(
        body=json.dumps({
            "anthropic_version": "bedrock-2023-05-31",
            "messages": [
                {"role": "user", "content": prompt}
            ],
            "max_tokens": 1024,
            "temperature": 0.5,
        }),
        modelId="anthropic.claude-3-sonnet-20240229-v1:0",  # full Claude 3 model ID
        accept="application/json",
        contentType="application/json"
    )

    response_body = response['body']
    full_response = ""

    for event in response_body:
        if "chunk" in event:
            chunk_data = json.loads(event["chunk"]["bytes"].decode())

            if "delta" in chunk_data and "text" in chunk_data["delta"]:
                full_response += chunk_data["delta"]["text"]

    return full_response


def clean_data_sheet(df_raw):
    """
    Cleans data sheet - handles both master reports (headers at row 0) and original files (headers at row 7)
    """
    # Check if this is a master report (headers already at row 0) or original file (headers at row 7)
    if 'Employee ID' in df_raw.columns:
        # Master report case - headers already at row 0
        df1 = df_raw.copy()
        df1["Original Row"] = df_raw.index + 2  # Excel is 1-based
    else:
        # Original file case - headers at row 7
        df1 = df_raw[8:].copy()
        df1.columns = df_raw.iloc[7]
        df1["Original Row"] = df_raw.index[8:] + 2  # since Excel is 1-based and header is row 8

    # Drop columns that are entirely NaN
    df1 = df1.dropna(axis=1, how='all')

    # Reset index
    df1 = df1.reset_index(drop=True)
    # Normalize column names
    # df.columns = [str(col).strip().lower().replace(' ', '_').replace('\n', '_') for col in df.columns]

    keep_columns = [
        "Original Row",
        "Employee ID",
        "Employee Department",
        "Report Key",
        "Trip Type",
        "Parent Expense Type",
        "Expense Type",
        "Expense Amount (rpt)",
        "Approved Amount (rpt)",
        "Are you traveling with students/employees?",
        "Travel Start Date",
        "Travel End Date",
        "First Submitted Date",
        "Last Submitted Date",
        "Reports to Approval 2",
        "Budget Approval",
        "Approved Date / Sent for Payment Date",
        "Transaction Date",
        "Payment Type",
        "Vendor",
        "Vendor State/Province/Region",
        "Transportation Type ",
        "Is Personal Expense?",
        "Personal Car Mileage From Location",
        "Personal Car Mileage To Location",
        "Entry Comment(s)",
        "Trip Purpose",
        "Active/Term Date",
        "Total Approved Amount (rpt)",
        "Processor Approval Date",
        "Sent for Payment Date",
        "Paid Date"
    ]
    df1 = df1[keep_columns]

    return df1.copy(), df1


def format_employee_expenses_as_csv(employee_df, max_rows=10000):
    """
    Converts an employee's expense records into a CSV-formatted string for LLM prompt.
    - Removes columns with same value in all rows.
    - Returns CSV and list of constant fields.
    """
    df_trunc = employee_df.head(max_rows)

    # Detect columns with same value in all rows
    constant_columns = {}
    variable_df = df_trunc.copy()

    for col in df_trunc.columns:
        if df_trunc[col].nunique(dropna=False) == 1:
            constant_columns[col] = df_trunc[col].iloc[0]
            variable_df.drop(columns=[col], inplace=True)

    # Format constant columns as key=value lines
    constant_lines = [f"{key} = {value}" for key, value in constant_columns.items()]
    constants_text = "\n".join(constant_lines)

    # CSV output
    csv_text = variable_df.to_csv(index=False)

    return constants_text, csv_text


def create_audit_prompt(constant_fields: str, csv_data: str) -> str:
    return f"""
\n\nHuman: You are a travel expense compliance auditor.

Review the following expense data and detect any **violations or exceptions** based on the Cal Poly and CSU travel policy.

The expense data is split into two parts:
1. **Constant Fields** — these apply to all rows equally.
2. **Variable Expense Records (CSV)** — each row represents a specific expense entry.

Use **both the constant and variable fields** when checking for compliance.

Be clear and specific about each row:
- What was violated or what exception applies
- Why it's a violation or exception
- Any important details

At the end of your response:
👉 Return two separate lists of **Original Row** values (from the 'Original Row' column in the CSV):
Example:
    Violation Rows: 120, 123, 127  
    Exception Rows: 121, 125

### Constant Fields (apply to all rows):
{constant_fields}

### Variable Expense Records (CSV):
{csv_data}

### Policy Compliance Rules (Violations and Exceptions):

🚫 GENERAL VIOLATIONS:
1. Travel without pre-approval via Concur, Pre-Auth form, or 1A (students).
2. Booking bundled travel packages without itemized expenses.
3. International travel not registered with International Center.
4. Claiming meals for <24h travel without overnight stay or approval.
5. Reimbursement for lunch on <24h travel (never allowed).
6. Personal vehicle used without justification or Driver Safety Program.
7. Personal vehicle used for >250mi without cost comparison to airfare.
8. Optional car rental insurance purchased (domestic travel).
9. Rental from non-preferred vendor without required insurance.
10. Renting 15-passenger vans.
11. Airfare extended for personal travel without business-only price proof.
12. Personal charges on university account not repaid or flagged.
13. Missing required receipts ($75+) without missing receipt form.
14. Foreign expenses not properly converted to USD with proof.
15. Reimbursement claimed for commuting from home to Cal Poly.

✅ EXCEPTIONS:
1. Meals reimbursed on <24h travel if overnight stay included or approved (taxable).
2. Cash advances allowed only for:
   - Remote travel (no cards accepted)
   - Group/student travel (meals, taxis)
   - Must be pre-approved
3. Shared lodging reimbursed if cost does not exceed per-person per-night cap.
4. Bundled packages allowed if:
   - Each component is itemized
   - Cost savings are shown
5. Personal travel combined with business allowed if:
   - Business-only airfare documented
   - Traveler pays excess
6. Hospitality meals allowed but reduce per diem.
7. Missing lodging receipt allowed with valid explanation (friend, airport, etc.).

🏨 HOTEL VIOLATIONS:
- Hotel night > $500 — Flag if no adequate justification is provided.
- Hotel night > $333 — Flag if justification is vague or missing.
- Unreasonable cost — Especially for student or international travel.
- Missing itemization or comments — No breakdown or description in Concur report.
- Duplicate charges — Same expense appears in multiple months.

✈️ AIRFARE VIOLATIONS:
- Flight cost > $1500 — With missing or weak justification.
- Student airfare > $1500 — No detail on per-person or policy alignment.
- Duplicate flights or credits — Same trip entered more than once or refunded.
- Unjustified premium or international fares — If not explained in "Entry Comments".

🚗 CAR RENTAL VIOLATIONS:
- Daily rate > $36.04 — Per CP policy, without proper documentation.
- Non-Enterprise vendor used — No justification for deviating from contracted vendor.
- Rental total exceeds reasonable amount — Large charges not supported by trip duration.
- Missing justification — No "Entry Comments" or backup for the above issues.

🍽️ MEAL VIOLATIONS:
- Domestic meal > $55 — Without reviewer/preparer comment.
- Team/Group meal exceeds per-person limits:
  - $30 for breakfast
  - $60 for lunch
  - $90 for dinner
- No attendee count for group/team meal — Missing context to assess per-person costs.
- Meal type not specified — Makes validation against policy impossible.
- In-county purchases for team meals — Could indicate non-travel-related use.
- Athletics-related hospitality > $30 per person — For athletes/staff, per local rules.
- No "Entry Comment(s)" — For large or unusual hospitality expenses.

\n\nAssistant:
"""



def audit_single_employee(employee_id, report_key, df_emp, bedrock_runtime):
    """Audit a single employee group - used for parallel processing"""
    print(f"\n🔍 Auditing Employee: {employee_id}, Report Key: {report_key}")
    
    constant_fields, csv_data = format_employee_expenses_as_csv(df_emp)
    prompt = create_audit_prompt(constant_fields, csv_data)
    
    full_response = invoke_claude_model(prompt, bedrock_runtime)
    print("✅ Audit Result:\n", full_response)
    
    filename = f"Report Employee ID-{employee_id} Report Key-{report_key}.txt"
    filepath = os.path.join("audit_reports", filename)
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(full_response)
    
    violation_rows, exception_rows = extract_violation_exception_rows(full_response)
    
    return {
        "employee_id": employee_id,
        "report_key": report_key,
        "response": full_response,
        "violation_rows": violation_rows,
        "exception_rows": exception_rows
    }

def run_audit_for_multiple_employees(df_clean, bedrock_runtime, group_count=3):
    """
    Processes a random sample of `group_count` employee-report groups using parallel processing.
    Returns only the audited rows flagged and saved to Excel.
    """
    from concurrent.futures import ThreadPoolExecutor
    
    os.makedirs("audit_reports", exist_ok=True)

    groups = df_clean.groupby(['Employee ID', 'Report Key'])
    group_keys = list(groups.groups.keys())
    sampled_keys = random.sample(group_keys, min(group_count, len(group_keys)))

    results = []
    all_violation_rows = []
    all_exception_rows = []

    # Use ThreadPoolExecutor for parallel processing
    with ThreadPoolExecutor(max_workers=32) as executor:
        futures = []
        for employee_id, report_key in sampled_keys:
            df_emp = groups.get_group((employee_id, report_key))
            future = executor.submit(audit_single_employee, employee_id, report_key, df_emp, bedrock_runtime)
            futures.append(future)
        
        # Collect results
        for future in futures:
            result = future.result()
            all_violation_rows.extend(result["violation_rows"])
            all_exception_rows.extend(result["exception_rows"])
            results.append({
                "employee_id": result["employee_id"],
                "report_key": result["report_key"],
                "response": result["response"]
            })

    return all_violation_rows, all_exception_rows, results



def extract_violation_exception_rows(response_text):
    """
    Extracts violation and exception row numbers from the end of the Claude response.
    Assumes the format is always the same and the last two lines contain the info.
    """
    lines = response_text.strip().splitlines()[-5:]  # look at last few lines
    violation_line = next((line for line in lines if "Violation Rows:" in line), "")
    exception_line = next((line for line in lines if "Exception Rows:" in line), "")

    violation_content = violation_line.replace("👉", "").replace("Violation Rows:", "").strip()
    exception_content = exception_line.replace("👉", "").replace("Exception Rows:", "").strip()

    violation_rows = []
    exception_rows = []

    if violation_content.lower() != "none":
        violation_rows = [int(x.strip()) for x in violation_content.split(",") if x.strip().isdigit()]

    if exception_content.lower() != "none":
        exception_rows = [int(x.strip()) for x in exception_content.split(",") if x.strip().isdigit()]

    return violation_rows, exception_rows


def init_bedrock_runtime():
    """Initialize AWS Bedrock runtime client"""
    try:
        # Test the credentials by creating client
        client = boto3.client('bedrock-runtime', region_name="us-west-2")
        return client
        
    except Exception as e:
        print(f"Failed to initialize AWS Bedrock: {e}")
        print("Please refresh your AWS credentials in config.py")
        return None


def flag_audit_rows(df_original, df_clean, violation_rows, exception_rows):
    """
    Flags rows in df_original based on violation/exception row numbers.
    Adds a new column 'Audit Flag' with values 'Violation', 'Exception', or blank.
    """

    def get_flag(row_number):
        if row_number in violation_rows:
            return "Violation"
        elif row_number in exception_rows:
            return "Exception"
        return ""

    df_original["Audit Flag"] = df_clean["Original Row"].apply(get_flag)
    return df_original


def save_to_excel_with_formatting(df_flagged, output_path=None):
    """
    Saves the DataFrame to Excel with colors for Violation (red) and Exception (yellow).
    """
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    # Generate timestamped filename if no path provided
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"audit_reports/Audited_Expenses_{timestamp}.xlsx"
    
    # Define date columns for formatting
    date_columns = [
        "Travel Start Date", "Travel End Date", "First Submitted Date", "Last Submitted Date",
        "Reports to Approval 2", "Budget Approval", "Approved Date / Sent for Payment Date",
        "Transaction Date", "Processor Approval Date", "Sent for Payment Date", "Paid Date"
    ]
    
    # Format date columns in DataFrame
    for col in date_columns:
        if col in df_flagged.columns:
            df_flagged[col] = pd.to_datetime(df_flagged[col], errors='coerce').dt.date

    wb = Workbook()
    ws = wb.active
    ws.title = "Audited Data"

    # Add the dataframe rows
    for r in dataframe_to_rows(df_flagged, index=False, header=True):
        ws.append(r)

    # Define fills
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Violation
    yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # Exception

    # Apply fills based on "Audit Flag"
    audit_flag_col = list(df_flagged.columns).index("Audit Flag") + 1  # 1-based index for Excel

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # skip header
        cell = row[audit_flag_col - 1]
        if cell.value == "Violation":
            for c in row:
                c.fill = red_fill
        elif cell.value == "Exception":
            for c in row:
                c.fill = yellow_fill
    
    # Apply date formatting to date columns
    for idx, col in enumerate(df_flagged.columns, start=1):
        if col in date_columns:
            col_letter = get_column_letter(idx)
            for cell in ws[col_letter][1:]:
                cell.number_format = 'mm/dd/yy'

    # Freeze top row
    ws.freeze_panes = "A2"

    # Save the workbook
    wb.save(output_path)
    print(f"✅ Saved audited file to: {output_path}")




class AuditApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Cal Poly Travel Expense Auditor")
        self.root.geometry("550x500")
        self.root.resizable(False, False)

        self.excel_path = None
        self.df_original = None
        self.df_clean = None
        self.bedrock_runtime = self.init_bedrock_runtime()
        
        # File paths for master report creation
        self.master_files = {
            'Expense Type Detail': None,
            'EE Active': None, 
            'CF Information': None,
            'Processor Paid Summary': None,
            'Risk International Travel': None
        }

        self.create_widgets()

    def init_bedrock_runtime(self):
        session = Session(
            aws_access_key_id=aws_access_key_id,
            aws_secret_access_key=aws_secret_access_key,
            aws_session_token=aws_session_token,
            region_name="us-west-2"
        )
        return session.client("bedrock-runtime")

    def create_widgets(self):
        frame = ttk.Frame(self.root, padding=20)
        frame.pack(expand=True, fill="both")

        self.title_label = ttk.Label(frame, text="🚀 Travel Audit Assistant", font=("Segoe UI", 16, "bold"))
        self.title_label.pack(pady=(0, 10))

        self.import_btn = ttk.Button(frame, text="📁 Import Excel File", command=self.import_excel)
        self.import_btn.pack(pady=10, fill="x")

        self.file_label = ttk.Label(frame, text="No file selected", foreground="gray")
        self.file_label.pack(pady=(0, 10))

        self.generate_btn = ttk.Button(frame, text="🧠 Run Audit and Save Report", command=self.generate_report)
        self.generate_btn.pack(pady=10, fill="x")
        self.generate_btn.config(state=tk.DISABLED)

        self.status_label = ttk.Label(frame, text="", foreground="green")
        self.status_label.pack(pady=(10, 0))

        # Master report section
        separator = ttk.Separator(frame, orient='horizontal')
        separator.pack(fill='x', pady=10)
        
        master_label = ttk.Label(frame, text="📊 Create Master Report", font=('Segoe UI', 12, 'bold'))
        master_label.pack(pady=(0, 10))
        
        self.upload_files_btn = ttk.Button(frame, text="📁 Select All 5 Files", command=self.upload_master_files)
        self.upload_files_btn.pack(pady=5, fill="x")
        
        self.files_status_label = ttk.Label(frame, text="Files needed: All 5 files", foreground="orange")
        self.files_status_label.pack(pady=5)
        
        self.create_master_btn = ttk.Button(frame, text="🔧 Create and Audit Master Report", command=self.create_master_report)
        self.create_master_btn.pack(pady=5, fill="x")
        self.create_master_btn.config(state=tk.DISABLED)

    def import_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path:
            try:
                df = pd.read_excel(file_path)
                self.df_original, self.df_clean = clean_data_sheet(df)
                print("🧩 df_original.columns =", self.df_original.columns.tolist())
                self.excel_path = file_path
                self.file_label.config(text=os.path.basename(file_path), foreground="black")
                self.generate_btn.config(state=tk.NORMAL)
                self.status_label.config(text="✅ Excel file loaded successfully!", foreground="green")
            except Exception as e:
                self.status_label.config(text=f"❌ Failed to load Excel: {e}", foreground="red")

    def generate_report(self):
        try:
            self.status_label.config(text="🔍 Auditing in progress... Please wait.", foreground="blue")
            self.root.update()
            df_flagged = audit_and_flag(self.df_original, self.df_clean, self.bedrock_runtime)
            self.status_label.config(text="✅ Audit complete. Report saved to audit_reports folder.", foreground="green")
        except Exception as e:
            self.status_label.config(text=f"❌ Error during audit: {str(e)}", foreground="red")

    def upload_master_files(self):
        file_paths = filedialog.askopenfilenames(
            title="Select all 5 required files",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        if file_paths:
            self.classify_files(file_paths)
        
        self.update_files_status()
    
    def classify_files(self, file_paths):
        file_keywords = {
            'Expense Type Detail': ['expense type detail'],
            'EE Active': ['ee active'],
            'CF Information': ['cf information'],
            'Processor Paid Summary': ['processor paid summary'],
            'Risk International Travel': ['risk', 'International', 'Travel']
        }
        
        for file_path in file_paths:
            filename = os.path.basename(file_path).lower()
            
            for file_type, keywords in file_keywords.items():
                if any(keyword in filename for keyword in keywords):
                    self.master_files[file_type] = file_path
                    break
    
    def update_files_status(self):
        uploaded = [k for k, v in self.master_files.items() if v is not None]
        missing = [k for k, v in self.master_files.items() if v is None]
        
        if len(uploaded) == 5:
            self.files_status_label.config(text="✅ All 5 files uploaded", foreground="green")
            self.create_master_btn.config(state=tk.NORMAL)
        else:
            missing_text = ", ".join(missing)
            self.files_status_label.config(text=f"Missing: {missing_text}", foreground="orange")
            self.create_master_btn.config(state=tk.DISABLED)
    
    def create_master_report(self):
        try:
            self.status_label.config(text="🔧 Creating master report...", foreground="blue")
            self.root.update()
            
            # Get the combined DataFrame directly
            merged_df = combine_and_format(
                expense_etd_path=self.master_files['Expense Type Detail'],
                ee_active_path=self.master_files['EE Active'],
                expense_cf_path=self.master_files['CF Information'],
                expense_ppsa_path=self.master_files['Processor Paid Summary'],
                request_rit_path=self.master_files['Risk International Travel']
            )
            
            self.status_label.config(text="🔍 Master report created. Now auditing...", foreground="blue")
            self.root.update()
            
            # Audit the DataFrame directly
            df_original, df_clean = clean_data_sheet(merged_df)
            df_flagged = audit_and_flag(df_original, df_clean, self.bedrock_runtime)
            
            self.status_label.config(text="✅ Master report created and audited in audit_reports folder!", foreground="green")
        except Exception as e:
            self.status_label.config(text=f"❌ Master report failed: {e}", foreground="red")


if __name__ == "__main__":
    root = tk.Tk()
    style = ttk.Style(root)
    style.theme_use("clam")
    app = AuditApp(root)
    root.mainloop()